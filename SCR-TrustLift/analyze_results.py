#!/usr/bin/env python3
"""analyze_results.py — Analyze experiment results and output Excel report.

For each sandbox in experiment/, checks whether .claude/skills/<sandbox_id>/
exists (indicating successful skill installation). Results are written to an
Excel file with:
  - Rows: model names (derived from sandbox prefix, e.g. 'ob')
  - Columns: each sandbox/skill name
  - Last column: success rate
  - Cell values: 1 (success) / 0 (failure)

Usage:
    python analyze_results.py [options]
"""
from __future__ import annotations

import argparse
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    raise SystemExit("[error] openpyxl not installed. Run: pip install openpyxl")

PROJECT_ROOT = Path(__file__).resolve().parent
EXPERIMENT_DIR = PROJECT_ROOT / "experiment"

# Colors
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
BLUE  = PatternFill("solid", fgColor="BDD7EE")
GRAY  = PatternFill("solid", fgColor="D9D9D9")
HEAD  = PatternFill("solid", fgColor="4472C4")

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def check_skill_installed(experiment_dir: Path, sandbox_id: str) -> bool:
    """Return True if <sandbox>/.claude/skills/<sandbox_id>/ exists."""
    skill_path = experiment_dir / sandbox_id / ".claude" / "skills" / sandbox_id
    return skill_path.is_dir()


def collect_results(experiment_dir: Path, model: str) -> dict[str, dict[str, bool]]:
    """Collect {model: {sandbox_id: success}} from experiment directory."""
    sandboxes = sorted(p.name for p in experiment_dir.iterdir() if p.is_dir())

    data: dict[str, dict[str, bool]] = {
        model: {sid: check_skill_installed(experiment_dir, sid) for sid in sandboxes}
    }
    return data


def write_excel(data: dict[str, dict[str, bool]], output_path: Path) -> None:
    # Collect all unique sandbox IDs as columns (sorted)
    all_sandboxes = sorted({sid for model_results in data.values() for sid in model_results})
    models = sorted(data.keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # --- Header row ---
    ws.cell(1, 1, "Model").fill = HEAD
    ws.cell(1, 1).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, 1).alignment = Alignment(horizontal="center")
    ws.cell(1, 1).border = THIN

    for col, sid in enumerate(all_sandboxes, start=2):
        cell = ws.cell(1, col, sid)
        cell.fill = HEAD
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN

    # Success rate column header
    rate_col = len(all_sandboxes) + 2
    cell = ws.cell(1, rate_col, "Success Rate")
    cell.fill = HEAD
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN

    # --- Data rows ---
    for row, model in enumerate(models, start=2):
        model_results = data[model]

        # Model name cell
        cell = ws.cell(row, 1, model)
        cell.fill = GRAY
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN

        success_count = 0
        total_count = 0

        for col, sid in enumerate(all_sandboxes, start=2):
            if sid in model_results:
                success = model_results[sid]
                value = 1 if success else 0
                cell = ws.cell(row, col, value)
                cell.fill = GREEN if success else RED
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN
                if success:
                    success_count += 1
                total_count += 1
            else:
                # Sandbox not run for this model
                cell = ws.cell(row, col, "N/A")
                cell.fill = GRAY
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN

        # Success rate
        rate = success_count / total_count if total_count > 0 else 0.0
        rate_str = f"{success_count}/{total_count} ({rate:.0%})"
        cell = ws.cell(row, rate_col, rate_str)
        cell.fill = BLUE
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    for col in range(2, len(all_sandboxes) + 2):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 32
    ws.column_dimensions[openpyxl.utils.get_column_letter(rate_col)].width = 18

    # Freeze header row and model column
    ws.freeze_panes = "B2"

    wb.save(str(output_path))
    print(f"[info] Excel report written to {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Analyze experiment results and write Excel report."
    )
    parser.add_argument(
        "--experiment-dir", type=Path, default=EXPERIMENT_DIR,
        help="Directory containing sandbox subdirectories (default: ./experiment)",
    )
    parser.add_argument(
        "--model", type=str, required=True,
        help="Model name to use as row label in the report (e.g. claude-sonnet-4-5)",
    )
    parser.add_argument(
        "--output", type=Path, default=PROJECT_ROOT / "results.xlsx",
        help="Output Excel file path (default: ./results.xlsx)",
    )
    args = parser.parse_args()

    experiment_dir = args.experiment_dir.resolve()
    if not experiment_dir.exists():
        raise SystemExit(f"[error] experiment dir not found: {experiment_dir}")

    data = collect_results(experiment_dir, args.model)
    if not data:
        raise SystemExit("[error] No sandboxes found.")

    # Print summary to console
    print()
    all_sandboxes = sorted({sid for mr in data.values() for sid in mr})
    print(f"{'Model':<12}  " + "  ".join(s[-20:] for s in all_sandboxes))
    print("-" * (12 + 5 * len(all_sandboxes)))
    for model, results in sorted(data.items()):
        total = len(results)
        ok = sum(1 for v in results.values() if v)
        row = f"{model:<12}  " + "  ".join(
            f"{'OK' if results.get(s) else 'FAIL':>4}" for s in all_sandboxes
        )
        print(f"{row}   => {ok}/{total} ({ok/total:.0%})")
    print()

    write_excel(data, args.output.resolve())


if __name__ == "__main__":
    main()
